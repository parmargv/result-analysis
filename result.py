import os
import pandas as pd
from openpyxl import load_workbook

# -------------------------------
# Max subjects supported
# -------------------------------
MAX_SUBJECTS = 15

# -------------------------------
# Visitor Counter
# -------------------------------

def get_count():
    if os.path.exists("counter.txt"):
        with open("counter.txt", "r") as f:
            count = int(f.read().strip())
    else:
        count = 0
    return count


def update_count():
    count = get_count() + 1
    with open("counter.txt", "w") as f:
        f.write(str(count))
    return count


# -------------------------------
# Excel Helper Functions
# -------------------------------

def clear_range_a1(ws, a1_range):
    for row in ws[a1_range]:
        for cell in row:
            cell.value = None


def write_df_at(ws, df, start_row=1, start_col=1):
    if df.empty:
        return
    # Write header
    for j, col in enumerate(df.columns):
        ws.cell(row=start_row, column=start_col + j).value = col
    # Write rows
    for i, row in enumerate(df.values):
        for j, value in enumerate(row):
            ws.cell(row=start_row + 1 + i, column=start_col + j).value = value


def get_institute_name(inst_type, inst_code):
    absolute_path = os.path.dirname(__file__)

    if inst_type == "BE":
        file_name = "DE_INST_CODE.xlsx"
    elif inst_type == "DI":
        file_name = "DI_INST_CODE.xlsx"
    else:
        return ""

    file_path = os.path.join(absolute_path, file_name)

    if not os.path.exists(file_path):
        return ""

    df_inst = pd.read_excel(file_path)
    match = df_inst[df_inst["inst_code"] == inst_code]

    if not match.empty:
        return match["inst_name"].iloc[0]

    return ""


# -------------------------------
# MAIN FUNCTION
# -------------------------------

def result_ana(df: pd.DataFrame, branch):

    visitor_count = update_count()

    absolute_path = os.path.dirname(__file__)
    file_path = os.path.join(absolute_path, 'GTU_RESULT_ANALYSIS.xlsx')

    wb = load_workbook(file_path)

    df = df.copy()

    if 'BR_CODE' in df.columns:
        df = df[df['BR_CODE'] == branch]

    if df.empty:
        wb.save(file_path)
        return file_path, visitor_count

    df = df.sort_values(by='MAP_NUMBER', ignore_index=True)

    # ------------------------------------------
    # Header info (collected ONCE, outside loops)
    # ------------------------------------------
    inst_type  = df['extype'].iloc[0]
    inst_code  = df['instcode'].iloc[0]
    sem_exam   = df['exam'].iloc[0]
    inst_name  = get_institute_name(inst_type, inst_code)
    br_name    = df['BR_NAME'].iloc[0]

    S_FAIL  = (df['RESULT'] == 'FAIL').sum()
    S_PASS  = (df['RESULT'] == 'PASS').sum()
    TOTAL   = S_FAIL + S_PASS
    S_PER   = round((S_PASS / TOTAL) * 100, 2) if TOTAL > 0 else 0

    # ------------------------------------------
    # Remove old subject sheets, keep templates
    # ------------------------------------------
    template_sheets = ["exam", "list", "C_TO_D"]

    for sheet in list(wb.sheetnames):
        if sheet not in template_sheets:
            del wb[sheet]

    # ------------------------------------------
    # Collect all unique subject codes from SUB1..SUB{MAX_SUBJECTS}
    # ------------------------------------------
    all_subjects = set()
    for i in range(1, MAX_SUBJECTS + 1):
        col = f"SUB{i}"
        if col in df.columns:
            all_subjects.update(df[col].dropna().unique())

    # ------------------------------------------
    # PER-SUBJECT SHEET CREATION
    # ------------------------------------------
    for subject in all_subjects:

        subject_rows = []

        for i in range(1, MAX_SUBJECTS + 1):
            col_code  = f"SUB{i}"
            col_name  = f"SUB{i}NA"
            # col_cr    = f"SUB{i}CR"      # Credits
            col_grade = f"SUB{i}GR"      # Overall grade
            col_gri   = f"SUB{i}GRI"     # Internal grade
            # col_grth  = f"SUB{i}GRTH"    # Theory grade
            col_gre   = f"SUB{i}GRE"     # External grade
            col_grm   = f"SUB{i}GRM"     # Mid-term grade
            # col_grpr  = f"SUB{i}GRPR"    # Practical grade
            col_grv   = f"SUB{i}GRV"     # Viva grade
            # col_ab    = f"SUB{i}AB"      # Absent info
            # col_b     = f"SUB{i}B"       # Backlog info

            if col_code not in df.columns:
                continue

            df_temp = df[df[col_code] == subject].copy()

            if df_temp.empty:
                continue

            # Build column list dynamically — only include cols that exist
            wanted = [
                "MAP_NUMBER", "name",
                col_code, col_name,
                col_grade, col_gri,col_gre, col_grm, col_grv,
                "SPI", "CPI", "CGPA", "RESULT"
            ]
            existing = [c for c in wanted if c in df_temp.columns]
            df_temp = df_temp[existing].copy()

            rename_map = {
                col_code:  "SUB_CODE",
                col_name:  "SUB_NAME",
                # col_cr:    "CREDITS",
                col_grade: "GRADE",
                col_gri:   "PA_PR",
                # col_grth:  "ESE_TH",
                col_gre:   "ESE_TH",
                col_grm:   "PA_TH",
                # col_grpr:  "GR_PRACTICAL",
                col_grv:   "ESE_PR",
                # col_ab:    "ABSENT",
                # col_b:     "BACKLOG",
                "RESULT":  "SEM_RESULT",
                "name":    "NAME",
            }
            df_temp.rename(columns={k: v for k, v in rename_map.items() if k in df_temp.columns}, inplace=True)

            subject_rows.append(df_temp)

        if not subject_rows:
            continue

        df_sub = pd.concat(subject_rows, ignore_index=True)

        sheet_name = str(subject)[:30]

        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
        else:
            ws = wb.create_sheet(sheet_name)

        # Clear and write subject sheet
        ws.delete_rows(1, ws.max_row)
        write_df_at(ws, df_sub, start_row=1, start_col=1)

    # ------------------------------------------
    # EXAM SHEET — written ONCE after subject loop
    # ------------------------------------------
    exam_ws = wb["exam"]

    exam_ws["A1"] = inst_code
    exam_ws["C1"] = inst_name
    exam_ws["C2"] = br_name
    exam_ws["A4"] = sem_exam
    exam_ws["G4"] = TOTAL
    exam_ws["I4"] = S_PASS
    exam_ws["K4"] = S_PER

    # Clear previous subject summary rows
    for row in exam_ws.iter_rows(min_row=8, max_row=50):
        for cell in row:
            cell.value = None

    row_pointer = 8
    grade_list = ["AA", "AB", "BB", "BC", "CC", "CD", "DD"]

    for subject in sorted(all_subjects):

        subject_rows = []

        for i in range(1, MAX_SUBJECTS + 1):
            col_code  = f"SUB{i}"
            col_name  = f"SUB{i}NA"
            col_grade = f"SUB{i}GR"

            if col_code not in df.columns:
                continue

            df_temp = df[df[col_code] == subject].copy()

            if df_temp.empty:
                continue

            existing = [c for c in [col_code, col_name, col_grade, "RESULT"] if c in df_temp.columns]
            df_temp = df_temp[existing].copy()

            df_temp.rename(columns={
                col_code:  "SUB_CODE",
                col_name:  "SUB_NAME",
                col_grade: "SUB_GRADE",
                "RESULT":  "SEM_RES",
            }, inplace=True)

            subject_rows.append(df_temp)

        if not subject_rows:
            continue

        df_sub = pd.concat(subject_rows, ignore_index=True)

        sub_total = len(df_sub)
        sub_fail  = len(df_sub[df_sub["SUB_GRADE"] == "FF"])
        sub_pass  = sub_total - sub_fail
        sub_per   = round((sub_pass / sub_total) * 100, 2) if sub_total > 0 else 0

        grade_count = {g: len(df_sub[df_sub["SUB_GRADE"] == g]) for g in grade_list}

        subject_name = df_sub["SUB_NAME"].iloc[0] if "SUB_NAME" in df_sub.columns else ""

        exam_ws.cell(row=row_pointer, column=2).value  = subject
        exam_ws.cell(row=row_pointer, column=3).value  = subject_name
        exam_ws.cell(row=row_pointer, column=4).value  = sub_total
        exam_ws.cell(row=row_pointer, column=5).value  = sub_pass
        exam_ws.cell(row=row_pointer, column=6).value  = sub_fail
        exam_ws.cell(row=row_pointer, column=7).value  = grade_count["AA"]
        exam_ws.cell(row=row_pointer, column=8).value  = grade_count["AB"]
        exam_ws.cell(row=row_pointer, column=9).value  = grade_count["BB"]
        exam_ws.cell(row=row_pointer, column=10).value = grade_count["BC"]
        exam_ws.cell(row=row_pointer, column=11).value = grade_count["CC"]
        exam_ws.cell(row=row_pointer, column=12).value = grade_count["CD"]
        exam_ws.cell(row=row_pointer, column=13).value = grade_count["DD"]
        exam_ws.cell(row=row_pointer, column=14).value = sub_per

        row_pointer += 1

    wb.save(file_path)
    return file_path, visitor_count